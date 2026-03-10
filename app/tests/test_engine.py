from __future__ import annotations

import math
import unittest
from pathlib import Path

from openpyxl import load_workbook

from app.domain.schemas import WinLoadInputSchema
from app.domain.service import build_response


WORKBOOK_PATH = Path(__file__).resolve().parents[3] / "wind_load_app" / "source_workbook.xlsm"


class EngineParityTests(unittest.TestCase):
    @classmethod
    def setUpClass(cls) -> None:
        cls.ws = load_workbook(WORKBOOK_PATH, data_only=True, keep_vba=True)["Win Load ASCE"]
        cls.response = build_response(WinLoadInputSchema())

    def assertCloseCell(self, cell: str, python_value: float) -> None:
        excel_value = float(self.ws[cell].value)
        self.assertTrue(
            math.isclose(excel_value, python_value, rel_tol=0.0, abs_tol=1e-12),
            msg=f"{cell}: excel={excel_value}, python={python_value}",
        )

    def test_intermediate_values_match_workbook(self) -> None:
        checks = {
            "M16": self.response["raw"]["design_speed_mps"],
            "M19": self.response["raw"]["roof_angle_deg"],
            "M20": self.response["raw"]["mean_roof_height_m"],
            "M21": self.response["raw"]["kz"],
            "M24": self.response["raw"]["qz_kn_per_m2"],
        }
        for cell, value in checks.items():
            with self.subTest(cell=cell):
                self.assertCloseCell(cell, float(value))

    def test_case_a_rows_match_workbook(self) -> None:
        first_row = self.response["raw"]["case_a_z1"]
        self.assertCloseCell("G34", first_row["gc_pf"])
        self.assertCloseCell("J34", first_row["wl1_or_we1"])
        self.assertCloseCell("M34", first_row["wl2_or_we2"])

    def test_line_load_labels_match_workbook(self) -> None:
        self.assertEqual(self.ws["S18"].value, self.response["outputs"]["line_load_labels"]["S18"])
        self.assertEqual(self.ws["AB41"].value, self.response["outputs"]["line_load_labels"]["AB41"])


if __name__ == "__main__":
    unittest.main()
